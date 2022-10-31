
import os.path
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

day = datetime.now()
month = day.strftime("%B_%Y")
where = 'nonhistory.xlsx'

def writer(name, reference, cheque_no, amount):
  print('writer activated')
  try:
    if os.path.exists(where):
      check = load_workbook(where)
      df = pd.DataFrame({'Name': [name],
                   'Reference': [reference],
                   'Cheque  Number': [cheque_no],
                   'Amount': [amount],
                   'Date': day.strftime("%d/%m/%Y %H:%M:%S")})
      if month in check.sheetnames:
        reader = pd.read_excel(where, sheet_name=month)
        writer = pd.ExcelWriter(where, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = check     
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        df.to_excel(writer,sheet_name=month,index=False,header=False,startrow=len(reader)+1)
      else:
        writer = pd.ExcelWriter(where, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = check 
        df.to_excel(writer, sheet_name=month, index=False)
      writer.close()
      return True
    else:
      print('inisialising')
      df = pd.DataFrame({'Name': [name],
                   'Reference': [reference],
                   'Cheque  Number': [cheque_no],
                   'Amount': [amount],
                   'Date': day.strftime("%d/%m/%Y %H:%M:%S")})
      writer = pd.ExcelWriter(where, engine='xlsxwriter')
      df.to_excel(writer, sheet_name=month, index=False)
      for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(str(column)))
        col_idx = df.columns.get_loc(column)
        writer.sheets[month].set_column(col_idx, col_idx, column_length)
      writer.save()
      return True
  except Exception as e:
    print('exception', e)
    return e

def reader():
  try:
    record = {}
    if day.month == 1:
      prew_month = datetime(day.year, 12, 6).strftime("%B_%Y")
    else:
      prew_month = datetime(day.year, day.month - 1, 6).strftime("%B_%Y")
    # Only to test if sheet exists
    check = load_workbook(where)
    if month in check.sheetnames:
      dfr = pd.read_excel(where, sheet_name=month)
      temp = dfr.values.tolist()
      temp.sort(key=lambda row: datetime.strptime(row[4], "%d/%m/%Y %H:%M:%S"), reverse=True)
      print('temp', dfr)
      record['month'] = temp
      record['last_num'] = int(temp[0][2]) + 1
    if prew_month in check.sheetnames:    
      dfr2 = pd.read_excel(where, sheet_name=prew_month)
      temp2 = dfr2.values.tolist()
      temp2.sort(key=lambda row: datetime.strptime(row[4], "%d/%m/%Y %H:%M:%S"), reverse=True)
      print('temp2', temp2)
      if 'last_num' not in record:
        record['last_num'] = int(temp2[0][2]) + 1
      record['prew_month'] = temp2
    return record
  except Exception as e:
    print('errpr', e)
    return None

def do_edit(cheque_date, check_num, name, reference, cheque_no, amount):
  print('starting find')
  try:
    d = datetime.strptime(cheque_date, "%d/%m/%Y %H:%M:%S")
    month = d.strftime("%B_%Y")
    print('at least here', check_num)
    check = load_workbook(where)
    if month in check.sheetnames:
      dfr = pd.read_excel(where, sheet_name=month)
      print('this far', dfr.index)
      for index in dfr.index:
        print('index', dfr.loc[index, 'Cheque  Number'])
        if str(dfr.loc[index, 'Cheque  Number']) == check_num:
          dfr.loc[index, 'Name'] = name
          dfr.loc[index, 'Reference'] = reference
          dfr.loc[index, 'Cheque  Number'] = cheque_no
          dfr.loc[index, 'Amount'] = amount
          print('dfr writting', dfr)
          writer = pd.ExcelWriter(where, engine='openpyxl', mode='a', if_sheet_exists="replace")
          dfr.to_excel(writer, sheet_name=month, index=False)
          writer.save()
          return ({'Success': True, 'msg': "Cheque for {} has been successfuly updated!".format(name)})
      return ({'Success': False, 'msg': "index not found"})
    else:
      return ({'Success': False, 'msg': "No entries found for {}".format(month)})
  except:
    return ({'Success': False, 'msg': "unknown error accured while searching for the entrie"})


def search(checque_num):
  start = datetime.now().timestamp()
  try:
    dfr = pd.read_excel(where, sheet_name=None)
    for key in dfr:
      out = dfr[key].loc[dfr[key]['Cheque  Number'] == checque_num].values.tolist()
      #print('test', dfr.iloc[np.where(dfr.A.values=='foo')])
      if out:
        finish = datetime.now().timestamp()
        print('time taken', finish - start)
        return out[0]
    finish = datetime.now().timestamp()
    print('time taken', finish - start)
    return False
  except Exception as e:
    print('errpr', e)

def stringify(n):
  if n == 0:
    return('Zero')
  elif n == 1:
    return('One')
  elif n == 2:
    return('Two')
  elif n == 3:
    return('Three')
  elif n == 4:
    return('Four')
  elif n == 5:
    return('Five')
  elif n == 6:
    return('Six')
  elif n == 7:
    return('Seven')
  elif n == 8:
    return('Eight')
  else:
    return('Nine')

def convert(amount):
  amount = float(amount)
  amount = int(amount)
  words = {}
  words['ten_mil'] = stringify(amount // 10000000)
  amount = amount % 10000000
  words['mill'] = stringify(amount // 1000000)
  amount = amount % 1000000
  words['hun'] = stringify(amount // 100000)
  amount = amount % 100000
  words['ten'] = stringify(amount // 10000)
  words['thousands'] = stringify(amount // 1000)
  amount = amount % 1000
  words['hundreds'] = stringify(amount // 100)
  amount = amount % 100
  words['tens'] = stringify(amount // 10)
  words['ones'] = stringify(amount % 10)
  return words